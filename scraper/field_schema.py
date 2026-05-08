from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

REQUIRED_FIELDS = ["name", "venuesName", "startDate", "startTime"]


def load_field_schema(xlsx_path: str | Path) -> list[str]:
    path = Path(xlsx_path)
    if not path.exists():
        return REQUIRED_FIELDS.copy()

    workbook = load_workbook(str(path), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [str(cell).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)) if cell]
    workbook.close()
    return headers or REQUIRED_FIELDS.copy()
